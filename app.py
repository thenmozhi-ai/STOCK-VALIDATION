"""
Multi-Marketplace Stock Validation App (Lazada / Shopee / TikTok / Zalora)
---------------------------------------------------------------------------
Upload StockValidation CSVs + marketplace stock files, get back a single
colour-coded Excel workbook with one tab-pair per marketplace plus a
combined summary dashboard.

Run locally:
    streamlit run app.py

Deploy on Streamlit Community Cloud:
    1. Push this folder to a GitHub repo.
    2. On share.streamlit.io, "New app" -> point at the repo / app.py.
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Multi-Marketplace Stock Validation", layout="wide")

# ----------------------------------------------------------------------------
# Constants / styling
# ----------------------------------------------------------------------------
NAVY = "1F3864"
GREEN_FILL, GREEN_FONT = "C6EFCE", "375623"
RED_FILL, RED_FONT = "FFC7CE", "9C0006"
ORANGE_FILL, ORANGE_FONT = "FFEB9C", "7D4800"
GREY_FILL, GREY_FONT = "D9D9D9", "595959"
ALT_ROW_FILL = "F2F7FB"
THIN_GREY = Side(style="thin", color="BFBFBF")

NOT_FOUND_LABELS = {
    "Lazada": "NOT IN S&P",
    "Shopee": "NOT IN SHOPEE",
    "TikTok": "NOT IN TIKTOK",
    "Zalora": "NOT IN ZALORA",
    "Shopify": "NOT IN SHOPIFY",
    "DTC": "NOT IN WAREHOUSE",
}

MARKETPLACE_ORDER = ["Lazada", "Shopee", "TikTok", "Zalora", "Shopify", "DTC"]

# Brand accent colours used purely for the Streamlit UI (cards / headers).
BRAND_COLORS = {
    "Lazada": "#FF6B35",
    "Shopee": "#F72585",
    "TikTok": "#4CC9F0",
    "Zalora": "#7209B7",
    "Shopify": "#95BF47",
    "DTC": "#3A86FF",
}
BRAND_ICONS = {
    "Lazada": "🛍️",
    "Shopee": "🧡",
    "TikTok": "🎵",
    "Zalora": "👗",
    "Shopify": "🟢",
    "DTC": "🏠",
}

# ----------------------------------------------------------------------------
# File-type detection
# ----------------------------------------------------------------------------
def classify_file(filename: str):
    """Return (marketplace, file_role) based on filename heuristics."""
    name = filename.lower()

    if "pricestock" in name:
        return "Lazada", "stock_price"
    if "mass_update_sales_info" in name:
        return "Shopee", "mass_update"
    if "tiktoksellercenter_batchedit" in name or "batchedit" in name:
        return "TikTok", "batch_edit"
    if "sellerstocktemplate" in name:
        return "Zalora", "stock_file"
    if "sellerstatustemplate" in name:
        return "Zalora", "status_file"
    if "sohbysku" in name:
        return "Warehouse", "soh"
    if name.startswith("all") and name.endswith(".csv"):
        return "Warehouse", "all_report"

    # StockValidation CSVs - identify marketplace by keyword in name
    if "stockvalidation" in name or name.endswith(".csv"):
        for mp in MARKETPLACE_ORDER:
            if mp.lower() in name:
                return mp, "stock_validation"

    return None, None


# ----------------------------------------------------------------------------
# Parsers per marketplace stock file
# ----------------------------------------------------------------------------
def fix_shopee_xlsx(file_bytes: bytes) -> bytes:
    """Patch the activePane XML bug present in Shopee's exported xlsx."""
    src = io.BytesIO(file_bytes)
    dst = io.BytesIO()
    with zipfile.ZipFile(src, "r") as zin, zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith(".xml") or item.filename.endswith(".rels"):
                text = data.decode("utf-8", errors="ignore")
                for old, new in [
                    ("bottom_left", "bottomLeft"),
                    ("top_left", "topLeft"),
                    ("bottom_right", "bottomRight"),
                    ("top_right", "topRight"),
                ]:
                    text = text.replace(f'activePane="{old}"', f'activePane="{new}"')
                    text = text.replace(f'pane="{old}"', f'pane="{new}"')
                data = text.encode("utf-8")
            zout.writestr(item, data)
    dst.seek(0)
    return dst.read()


def parse_lazada_stock_price(file_bytes: bytes) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes), header=0, skiprows=3)
    df = df.iloc[:, :15]
    df.columns = [
        "Product ID", "catId", "Product Name", "currencyCode", "sku.skuId",
        "status", "Shop SKU", "SellerSKU", "Quantity", "Price", "SpecialPrice",
        "SpecialPrice Start", "SpecialPrice End", "Variations Combo", "md5key",
    ]
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("SellerSKU")["Quantity"].sum().to_dict()


def parse_shopee_mass_update(file_bytes: bytes) -> dict:
    fixed = fix_shopee_xlsx(file_bytes)
    df = pd.read_excel(io.BytesIO(fixed), header=2, skiprows=[3, 4])
    df = df.iloc[:, :10]
    df.columns = [
        "Product ID", "Category", "Product Name", "Parent SKU", "SKU",
        "Price", "GTIN", "Quantity", "Minimum Purchase Quantity", "Fail Reason",
    ]
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("SKU")["Quantity"].sum().to_dict()


def parse_tiktok_batch_edit(file_bytes: bytes) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes), header=2, skiprows=[3, 4])
    df = df.iloc[:, :8]
    df.columns = [
        "Product ID", "Category", "Product Name", "SKU ID",
        "Variation Option", "Price", "Quantity", "Seller SKU",
    ]
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("Seller SKU")["Quantity"].sum().to_dict()


def parse_zalora_stock_file(file_bytes: bytes) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    df = df.iloc[:, :4]
    df.columns = ["SellerSku", "ShopSku", "Quantity", "Name"]
    df["SellerSku"] = df["SellerSku"].astype(str).str.strip()
    df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0).astype(int)
    return df.groupby("SellerSku")["Quantity"].sum().to_dict()


def parse_zalora_status_file(file_bytes: bytes) -> dict:
    df = pd.read_excel(io.BytesIO(file_bytes), header=0)
    df = df.iloc[:, :4]
    df.columns = ["SellerSku", "ShopSku", "Name", "Status"]
    df["SellerSku"] = df["SellerSku"].astype(str).str.strip()
    return df.set_index("SellerSku")["Status"].astype(str).str.strip().str.lower().to_dict()


def parse_shopify_export(file_bytes: bytes, filename: str) -> dict:
    """Standard Shopify 'Export inventory' CSV. Auto-detects the SKU column
    and the best available quantity column (Available > On hand > Quantity)."""
    if filename.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]

    sku_col = next((c for c in df.columns if c.lower() == "sku"), None) \
        or next((c for c in df.columns if "sku" in c.lower()), None)
    qty_col = next((c for c in df.columns if c.lower() == "available"), None) \
        or next((c for c in df.columns if "on hand" in c.lower()), None) \
        or next((c for c in df.columns if "quantity" in c.lower()), None)

    if not sku_col or not qty_col:
        return {}

    df[sku_col] = df[sku_col].astype(str).str.strip()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)
    return df.groupby(sku_col)[qty_col].sum().to_dict()


def parse_soh(file_bytes: bytes) -> dict:
    """SOHbySKU is an XML-formatted .xls file. Data rows start at index 7."""
    ns = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}
    root = ET.fromstring(file_bytes)
    rows = root.findall(".//ss:Row", ns)
    lookup = {}
    for row in rows[7:]:
        cells = row.findall("ss:Cell", ns)
        values = []
        for cell in cells:
            data_el = cell.find("ss:Data", ns)
            values.append(data_el.text if data_el is not None else None)
        if len(values) > 14 and values[6]:
            sku = str(values[6]).strip()
            try:
                qty = int(float(values[14]))
            except (TypeError, ValueError):
                qty = 0
            lookup[sku] = lookup.get(sku, 0) + qty
    return lookup


def parse_warehouse_report(file_bytes: bytes, filename: str) -> dict:
    """Generic SellerSKU + quantity warehouse report from any CSV/XLSX source.
    Auto-detects a SKU-like column and a quantity-like column."""
    name = filename.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    elif name.endswith(".xls"):
        # Some warehouse exports still use the legacy XML-XLS format like SOHbySKU.
        try:
            return parse_soh(file_bytes)
        except Exception:
            df = pd.read_excel(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))

    df.columns = [str(c).strip() for c in df.columns]
    sku_col = next((c for c in df.columns if "sku" in c.lower()), None)
    qty_col = next((c for c in df.columns if "qty" in c.lower() or "quantity" in c.lower()
                     or "stock" in c.lower()), None)
    if not sku_col or not qty_col:
        return {}

    df[sku_col] = df[sku_col].astype(str).str.strip()
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)
    return df.groupby(sku_col)[qty_col].sum().to_dict()


def build_warehouse_vs_soh_df(warehouse_lookup: dict, soh_lookup: dict) -> pd.DataFrame:
    """Compare a Warehouse report's stock against SOH, SKU by SKU."""
    all_skus = set(warehouse_lookup) | set(soh_lookup)
    rows = []
    for sku in sorted(all_skus):
        wh_qty = warehouse_lookup.get(sku)
        soh_qty = soh_lookup.get(sku)
        if wh_qty is None:
            continue
        status, remark = get_status_remark(wh_qty, soh_qty, "NOT IN SOH")
        rows.append({"Seller SKU": sku, "Expected Stock": wh_qty, "SP_Quantity": soh_qty,
                     "Status": status, "Remark": remark})
    return pd.DataFrame(rows)


def parse_all_report(file_bytes: bytes) -> dict:
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]
    qty_col = df.columns[14] if len(df.columns) > 14 else None
    if "sellerSKU" not in df.columns or qty_col is None:
        return {}
    df[qty_col] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0).astype(int)
    return df.groupby("sellerSKU")[qty_col].sum().to_dict()


def parse_stock_validation_csv(file_bytes: bytes) -> pd.DataFrame:
    df = pd.read_csv(io.BytesIO(file_bytes))
    df.columns = [c.strip() for c in df.columns]
    df["Expected Stock"] = pd.to_numeric(df["Expected Stock"], errors="coerce").fillna(0).astype(int)
    df["Seller SKU"] = df["Seller SKU"].astype(str).str.strip()
    return df


# ----------------------------------------------------------------------------
# Remark logic (identical across marketplaces)
# ----------------------------------------------------------------------------
def get_status_remark(expected, sp_quantity, not_found_label):
    if pd.isna(sp_quantity):
        return "Mismatch", not_found_label
    exp = int(expected)
    sp = int(sp_quantity)
    if exp == 0 and sp > 0:
        return "Mismatch", "UPDATE 0"
    if exp != sp:
        return "Mismatch", "IMPACT"
    return "Match", "TRUE"


def build_marketplace_df(stockval_df: pd.DataFrame, sp_lookup: dict, marketplace: str,
                          status_lookup: dict = None):
    df = stockval_df.copy()
    df["SP_Quantity"] = df["Seller SKU"].map(sp_lookup)
    not_found_label = NOT_FOUND_LABELS[marketplace]

    statuses, remarks = [], []
    for _, row in df.iterrows():
        status, remark = get_status_remark(row["Expected Stock"], row["SP_Quantity"], not_found_label)
        statuses.append(status)
        remarks.append(remark)
    df["Status"] = statuses
    df["Remark"] = remarks

    if marketplace == "Zalora" and status_lookup:
        df["Zalora_Status"] = df["Seller SKU"].map(status_lookup).fillna("—")
        cols = list(df.columns)
        cols.remove("Zalora_Status")
        insert_at = cols.index("SP_Quantity") + 1
        cols.insert(insert_at, "Zalora_Status")
        df = df[cols]

    return df


# ----------------------------------------------------------------------------
# Excel writer
# ----------------------------------------------------------------------------
def style_header(ws, ncols, row=1, height=36):
    ws.row_dimensions[row].height = height
    fill = PatternFill("solid", fgColor=NAVY)
    font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=THIN_GREY)


def write_df_sheet(wb, sheet_name, df, remark_col="Remark", status_col="Status"):
    ws = wb.create_sheet(sheet_name[:31])
    ws.append(list(df.columns))
    style_header(ws, len(df.columns))

    remark_idx = df.columns.get_loc(remark_col) + 1 if remark_col in df.columns else None
    status_idx = df.columns.get_loc(status_col) + 1 if status_col in df.columns else None
    pm_check_idx = df.columns.get_loc("In Product Master") + 1 if "In Product Master" in df.columns else None

    color_map = {
        "TRUE": (GREEN_FILL, GREEN_FONT),
        "Match": (GREEN_FILL, GREEN_FONT),
        "FOUND IN PM": (GREEN_FILL, GREEN_FONT),
        "IMPACT": (RED_FILL, RED_FONT),
        "Mismatch": (RED_FILL, RED_FONT),
        "UPDATE 0": (ORANGE_FILL, ORANGE_FONT),
        "NOT IN PRODUCT MASTER": (GREY_FILL, GREY_FONT),
        "NOT IN SOH": (GREY_FILL, GREY_FONT),
    }
    for label in NOT_FOUND_LABELS.values():
        color_map[label] = (GREY_FILL, GREY_FONT)

    for r_i, row in enumerate(df.itertuples(index=False), start=2):
        for c_i, value in enumerate(row, start=1):
            cell = ws.cell(row=r_i, column=c_i, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.border = Border(top=THIN_GREY, bottom=THIN_GREY, left=THIN_GREY, right=THIN_GREY)
            if r_i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW_FILL)
        remark_val = df.iloc[r_i - 2][remark_col] if remark_col in df.columns else None
        fill_color, font_color = color_map.get(remark_val, (None, None))
        if fill_color:
            for idx in (remark_idx, status_idx):
                if idx:
                    cell = ws.cell(row=r_i, column=idx)
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = Font(name="Arial", size=10, bold=True, color=font_color)

        if pm_check_idx:
            pm_val = df.iloc[r_i - 2]["In Product Master"]
            pm_fill, pm_font = (GREEN_FILL, GREEN_FONT) if pm_val == "Yes" else (GREY_FILL, GREY_FONT)
            cell = ws.cell(row=r_i, column=pm_check_idx)
            cell.fill = PatternFill("solid", fgColor=pm_fill)
            cell.font = Font(name="Arial", size=10, bold=True, color=pm_font)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for i, col in enumerate(df.columns, start=1):
        if len(df):
            max_len = df[col].astype(str).str.len().max()
            max_len = 10 if pd.isna(max_len) else int(max_len)
        else:
            max_len = 10
        width = max(12, min(30, max_len + 4))
        if col == "Zalora_Status":
            width = 16
        ws.column_dimensions[get_column_letter(i)].width = width
    return ws


def add_summary_block(ws, start_row, title, counts, labels=None):
    fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(row=start_row, column=1, value=title).font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    ws.cell(row=start_row, column=1).fill = fill
    for c in range(2, 7):
        ws.cell(row=start_row, column=c).fill = fill

    labels = labels or ["Total", "TRUE", "IMPACT", "UPDATE 0", "NOT FOUND", "Total Mismatches"]
    colors = [None, (GREEN_FILL, GREEN_FONT), (RED_FILL, RED_FONT), (ORANGE_FILL, ORANGE_FONT),
              (GREY_FILL, GREY_FONT), (RED_FILL, RED_FONT)]
    values = [
        counts["total"], counts["true"], counts["impact"],
        counts["update0"], counts["not_found"], counts["mismatches"],
    ]
    for i, (label, val, color) in enumerate(zip(labels, values, colors)):
        r = start_row + 1
        c = i + 1
        ws.cell(row=r, column=c, value=label).font = Font(name="Arial", size=10, bold=True)
        cell = ws.cell(row=r + 1, column=c, value=val)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center")
        if color:
            cell.fill = PatternFill("solid", fgColor=color[0])
            cell.font = Font(name="Arial", size=11, bold=True, color=color[1])
    return start_row + 4


def compute_counts(df, not_found_label, match_label="TRUE"):
    total = len(df)
    true_ct = (df["Remark"] == match_label).sum()
    impact_ct = (df["Remark"] == "IMPACT").sum()
    update0_ct = (df["Remark"] == "UPDATE 0").sum()
    nf_ct = (df["Remark"] == not_found_label).sum()
    mismatches = (df["Status"] == "Mismatch").sum()
    return dict(total=total, true=true_ct, impact=impact_ct, update0=update0_ct,
                not_found=nf_ct, mismatches=mismatches)


def build_workbook(marketplace_data: dict, warehouse_df: pd.DataFrame = None, brand_name="Shop",
                    mp_vs_pm_df: pd.DataFrame = None, warehouse_vs_soh_df: pd.DataFrame = None):
    wb = Workbook()
    wb.remove(wb.active)
    summary_ws = wb.create_sheet("Summary")
    summary_ws.column_dimensions["A"].width = 20
    for col in "BCDEF":
        summary_ws.column_dimensions[col].width = 16

    summary_ws.cell(row=1, column=1,
                     value=f"📊  Multi-Marketplace Stock Validation — {brand_name}").font = Font(
        name="Arial", size=14, bold=True)
    row_cursor = 3

    for marketplace in MARKETPLACE_ORDER:
        if marketplace not in marketplace_data:
            continue
        df = marketplace_data[marketplace]
        not_found_label = NOT_FOUND_LABELS[marketplace]
        counts = compute_counts(df, not_found_label)
        row_cursor = add_summary_block(summary_ws, row_cursor, marketplace.upper(), counts)

        write_df_sheet(wb, f"{marketplace} - StockVal", df)
        mism_df = df[df["Status"] == "Mismatch"]
        write_df_sheet(wb, f"{marketplace} - Mismatches", mism_df)

    if warehouse_df is not None:
        counts = compute_counts(warehouse_df, "NOT FOUND")
        row_cursor = add_summary_block(summary_ws, row_cursor, "SOH (Reference)", counts)
        write_df_sheet(wb, "SOH Reference", warehouse_df)

    if mp_vs_pm_df is not None:
        counts = compute_counts(mp_vs_pm_df, "NOT IN PRODUCT MASTER", match_label="FOUND IN PM")
        row_cursor = add_summary_block(
            summary_ws, row_cursor, "MP vs PRODUCT MASTER", counts,
            labels=["Total", "Found in PM", "—", "—", "Not in PM", "Total Not Found"],
        )
        write_df_sheet(wb, "MP vs Product Master", mp_vs_pm_df)
        mism_df = mp_vs_pm_df[mp_vs_pm_df["Status"] == "Mismatch"]
        write_df_sheet(wb, "MP vs PM Mismatches", mism_df)

    if warehouse_vs_soh_df is not None:
        counts = compute_counts(warehouse_vs_soh_df, "NOT IN SOH")
        row_cursor = add_summary_block(summary_ws, row_cursor, "WAREHOUSE vs SOH", counts)
        write_df_sheet(wb, "Warehouse vs SOH", warehouse_vs_soh_df)
        mism_df = warehouse_vs_soh_df[warehouse_vs_soh_df["Status"] == "Mismatch"]
        write_df_sheet(wb, "Warehouse vs SOH Mismatches", mism_df)

    return wb


def parse_product_master(file_bytes: bytes, filename: str) -> dict:
    """Best-effort SKU -> Product Name lookup from a master product file.
    Auto-detects a SKU-like column and a Name-like column by header keyword."""
    try:
        if filename.lower().endswith(".csv"):
            df = pd.read_csv(io.BytesIO(file_bytes))
        else:
            df = pd.read_excel(io.BytesIO(file_bytes))
    except Exception:
        return {}

    df.columns = [str(c).strip() for c in df.columns]
    sku_col = next((c for c in df.columns if "sku" in c.lower()), None)
    name_col = next((c for c in df.columns if "name" in c.lower() or "product" in c.lower()), None)
    if not sku_col or not name_col:
        return {}

    df[sku_col] = df[sku_col].astype(str).str.strip()
    return df.set_index(sku_col)[name_col].to_dict()


def apply_product_names(df: pd.DataFrame, name_lookup: dict) -> pd.DataFrame:
    if not name_lookup or "Seller SKU" not in df.columns:
        return df
    df = df.copy()
    df["Product Name"] = df["Seller SKU"].map(name_lookup).fillna("—")
    cols = list(df.columns)
    cols.remove("Product Name")
    insert_at = cols.index("Seller SKU") + 1
    cols.insert(insert_at, "Product Name")
    return df[cols]


def apply_product_master_check(df: pd.DataFrame, name_lookup: dict) -> pd.DataFrame:
    """Adds 'Product Name' + an explicit 'In Product Master' Yes/No cross-check column,
    matched by Seller SKU. Used for Warehouse report and DTC report tabs."""
    df = apply_product_names(df, name_lookup)
    if not name_lookup or "Seller SKU" not in df.columns:
        return df
    df = df.copy()
    df["In Product Master"] = df["Seller SKU"].isin(name_lookup.keys()).map({True: "Yes", False: "No"})
    cols = list(df.columns)
    cols.remove("In Product Master")
    insert_at = cols.index("Product Name") + 1 if "Product Name" in cols else cols.index("Seller SKU") + 1
    cols.insert(insert_at, "In Product Master")
    return df[cols]


def parse_mp_report(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Generic MP report reader — auto-detects a SKU column and (optionally)
    a quantity-like column. Returns a standardised dataframe."""
    if filename.lower().endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes))
    else:
        df = pd.read_excel(io.BytesIO(file_bytes))
    df.columns = [str(c).strip() for c in df.columns]

    sku_col = next((c for c in df.columns if "sku" in c.lower()), None)
    if not sku_col:
        return pd.DataFrame()

    qty_col = next((c for c in df.columns if "qty" in c.lower() or "quantity" in c.lower()
                     or "stock" in c.lower()), None)

    out = pd.DataFrame()
    out["Seller SKU"] = df[sku_col].astype(str).str.strip()
    out = out.drop_duplicates(subset="Seller SKU")
    if qty_col:
        qty_lookup = df.assign(**{sku_col: df[sku_col].astype(str).str.strip()}) \
            .groupby(sku_col)[qty_col].sum()
        out["MP_Quantity"] = out["Seller SKU"].map(qty_lookup)
    return out


def build_mp_vs_product_master_df(mp_df: pd.DataFrame, name_lookup: dict) -> pd.DataFrame:
    """Compare an MP report's SKUs against the Product Master SKU list."""
    df = mp_df.copy()
    df["Product Name"] = df["Seller SKU"].map(name_lookup)
    df["Status"] = df["Product Name"].apply(lambda x: "Match" if pd.notna(x) else "Mismatch")
    df["Remark"] = df["Product Name"].apply(lambda x: "FOUND IN PM" if pd.notna(x) else "NOT IN PRODUCT MASTER")
    df["Product Name"] = df["Product Name"].fillna("—")
    cols = list(df.columns)
    cols.remove("Product Name")
    insert_at = cols.index("Seller SKU") + 1
    cols.insert(insert_at, "Product Name")
    return df[cols]


# ----------------------------------------------------------------------------
# Streamlit UI
# ----------------------------------------------------------------------------
st.markdown(
    """
    <style>
    .main .block-container { padding-top: 2rem; }
    .mp-card {
        border-radius: 12px;
        padding: 1.1rem 1.3rem 1.3rem 1.3rem;
        margin-bottom: 1.2rem;
        border: 1px solid rgba(0,0,0,0.08);
        box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .mp-card-header {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        font-size: 1.15rem;
        font-weight: 700;
        margin-bottom: 0.6rem;
        color: white;
        padding: 0.45rem 0.9rem;
        border-radius: 8px;
    }
    .ref-card {
        border-radius: 12px;
        padding: 1.1rem 1.3rem 1.3rem 1.3rem;
        margin-bottom: 1.2rem;
        border: 1px solid rgba(0,0,0,0.08);
        background: linear-gradient(135deg, #f5f7fa 0%, #eef1f6 100%);
    }
    .legend-chip {
        display: inline-block;
        padding: 0.25rem 0.7rem;
        border-radius: 14px;
        font-size: 0.82rem;
        font-weight: 600;
        margin-right: 0.5rem;
    }
    .status-badge-ok { color: #1a7f37; font-weight: 700; }
    .status-badge-missing { color: #b3261e; font-weight: 700; }
    div.stButton > button[kind="primary"] {
        background: linear-gradient(90deg, #FF6B35 0%, #F72585 50%, #7209B7 100%);
        border: none;
        font-weight: 700;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def mp_card_header(marketplace: str):
    color = BRAND_COLORS[marketplace]
    icon = BRAND_ICONS[marketplace]
    st.markdown(
        f'<div class="mp-card-header" style="background:{color};">{icon} {marketplace}</div>',
        unsafe_allow_html=True,
    )


def readiness_line(mp_file, inv_file, extra_ok=True):
    if mp_file and inv_file and extra_ok:
        st.markdown('<span class="status-badge-ok">✅ Ready to validate</span>', unsafe_allow_html=True)
    elif mp_file or inv_file:
        missing = []
        if not mp_file:
            missing.append("MP file")
        if not inv_file:
            missing.append("inventory file")
        st.markdown(
            f'<span class="status-badge-missing">⚠️ Missing: {", ".join(missing)}</span>',
            unsafe_allow_html=True,
        )
    else:
        st.caption("Not uploaded — this marketplace will be skipped.")


st.title("📊 Multi-Marketplace Stock Validation")
st.caption("Lazada · Shopee · TikTok · Zalora · Shopify — upload each pair of files, get one colour-coded workbook.")

with st.container():
    st.markdown(
        '<span class="legend-chip" style="background:#C6EFCE;color:#375623;">✅ TRUE / Match</span>'
        '<span class="legend-chip" style="background:#FFC7CE;color:#9C0006;">❌ IMPACT / Mismatch</span>'
        '<span class="legend-chip" style="background:#FFEB9C;color:#7D4800;">🟠 UPDATE 0</span>'
        '<span class="legend-chip" style="background:#D9D9D9;color:#595959;">⬜ NOT FOUND</span>',
        unsafe_allow_html=True,
    )

st.write("")
brand_name = st.text_input("Brand / Shop name (shown on the summary tab)", value="My Shop")

st.markdown("### 1️⃣ Marketplace pairs")
st.caption("Each marketplace needs BOTH its MP file and its inventory (StockValidation) file to run.")

row1 = st.columns(3)
row2 = st.columns(3)

with row1[0]:
    with st.container(border=True):
        mp_card_header("Lazada")
        lazada_mp_file = st.file_uploader("Lazada MP file", type=["xlsx"], key="lazada_mp")
        lazada_inv_file = st.file_uploader("Lazada inventory file", type=["csv"], key="lazada_inv")
        readiness_line(lazada_mp_file, lazada_inv_file)

with row1[1]:
    with st.container(border=True):
        mp_card_header("Shopee")
        shopee_mp_file = st.file_uploader("Shopee MP file", type=["xlsx"], key="shopee_mp")
        tc_shopee_inv_file = st.file_uploader("TC Shopee inventory file", type=["csv"], key="tc_shopee_inv")
        readiness_line(shopee_mp_file, tc_shopee_inv_file)

with row1[2]:
    with st.container(border=True):
        mp_card_header("TikTok")
        tiktok_mp_file = st.file_uploader("Tiktok MP file", type=["xlsx"], key="tiktok_mp")
        tiktok_inv_file = st.file_uploader("Tiktok inventory file", type=["csv"], key="tiktok_inv")
        readiness_line(tiktok_mp_file, tiktok_inv_file)

with row2[0]:
    with st.container(border=True):
        mp_card_header("Zalora")
        zalora_mp_file = st.file_uploader("Zalora MP file (SellerStockTemplate)", type=["xlsx"], key="zalora_mp")
        zalora_inv_file = st.file_uploader("Zalora inventory file", type=["csv"], key="zalora_inv")
        zalora_status_file = st.file_uploader(
            "Zalora Status file (optional, SellerStatusTemplate)", type=["xlsx"], key="zalora_status")
        readiness_line(zalora_mp_file, zalora_inv_file)

with row2[1]:
    with st.container(border=True):
        mp_card_header("Shopify")
        shopify_mp_file = st.file_uploader("Shopify MP file (Export inventory CSV)", type=["csv", "xlsx"], key="shopify_mp")
        shopify_inv_file = st.file_uploader("Shopify inventory file", type=["csv"], key="shopify_inv")
        readiness_line(shopify_mp_file, shopify_inv_file)

with row2[2]:
    with st.container(border=True):
        mp_card_header("DTC")
        st.caption("Compares against the SOH warehouse file below — no separate MP file needed here.")
        dtc_inv_file = st.file_uploader("DTC inventory file", type=["csv"], key="dtc_inv")
        dtc_readiness_slot = st.empty()

st.markdown("### 2️⃣ Warehouse & reference (optional)")
with st.container(border=True):
    st.markdown('<div class="mp-card-header" style="background:#1F3864;">📦 Warehouse & Product Master</div>',
                unsafe_allow_html=True)
    wh_col1, wh_col2, wh_col3, wh_col4 = st.columns(4)
    with wh_col1:
        soh_file = st.file_uploader("SOH", type=["xls"], key="soh")
    with wh_col2:
        warehouse_report_file = st.file_uploader("Warehouse report", type=["csv", "xlsx"], key="warehouse_report")
        st.caption("Compared against SOH by SKU.")
    with wh_col3:
        product_master_file = st.file_uploader("Product Master file", type=["csv", "xlsx"], key="product_master")
    with wh_col4:
        mp_report_file = st.file_uploader("MP Report file", type=["csv", "xlsx"], key="mp_report")
        st.caption("Checked against Product Master by SKU — needs Product Master uploaded too.")

with dtc_readiness_slot:
    readiness_line(soh_file, dtc_inv_file)

with st.expander("ℹ️ Notes on file formats", expanded=False):
    st.markdown(
        """
- **Lazada MP file** — the `pricestock...xlsx` Stock & Price export
- **Shopee MP file** — the `mass_update_sales_info...xlsx` export
- **Tiktok MP file** — the `Tiktoksellercenter_batchedit...xlsx` export
- **Zalora MP file** — `SellerStockTemplate...xlsx`; optional Status file adds an active/inactive column
- **Shopify MP file** — the standard Shopify "Export inventory" CSV (`SKU` + `Available` columns)
- **DTC inventory file** — StockValidation-style CSV for your own DTC site; validated against the
  SOH warehouse file (upload that too)
- **Warehouse report** — a second warehouse stock source (SKU + quantity, any column names);
  compared directly against SOH, SKU by SKU
- **Inventory files** — the StockValidation CSV for that marketplace (`Seller SKU` + `Expected Stock` columns)
- **SOH** — `SOHbySKU...xls` warehouse export (shown as reference only in this version)
- **Product Master file** — any file with a SKU column and a Name column; adds a `Product Name`
  column to marketplace reports, doesn't affect matching
- **MP Report file** — any file with a SKU column; checked against the Product Master's SKU list
  to flag which SKUs are missing from the master

Only fill in the marketplaces you want validated — any subset works.
        """
    )

any_uploaded = any([
    lazada_mp_file, shopee_mp_file, tiktok_mp_file, zalora_mp_file, shopify_mp_file, dtc_inv_file,
    lazada_inv_file, tc_shopee_inv_file, tiktok_inv_file, zalora_inv_file, shopify_inv_file,
    soh_file, product_master_file, mp_report_file, warehouse_report_file,
])

if any_uploaded:
    name_lookup = {}
    if product_master_file:
        name_lookup = parse_product_master(product_master_file.read(), product_master_file.name)
        if name_lookup:
            st.success(f"Product Master loaded — {len(name_lookup)} SKUs mapped to product names.")
        else:
            st.warning("Product Master file uploaded but no SKU/Name columns could be detected — skipping.")

    if st.button("🚀 Run Validation", type="primary"):
        marketplace_data = {}

        if lazada_inv_file and lazada_mp_file:
            stockval_df = parse_stock_validation_csv(lazada_inv_file.read())
            sp_lookup = parse_lazada_stock_price(lazada_mp_file.read())
            df = build_marketplace_df(stockval_df, sp_lookup, "Lazada")
            marketplace_data["Lazada"] = apply_product_names(df, name_lookup)

        if tc_shopee_inv_file and shopee_mp_file:
            stockval_df = parse_stock_validation_csv(tc_shopee_inv_file.read())
            sp_lookup = parse_shopee_mass_update(shopee_mp_file.read())
            df = build_marketplace_df(stockval_df, sp_lookup, "Shopee")
            marketplace_data["Shopee"] = apply_product_names(df, name_lookup)

        if tiktok_inv_file and tiktok_mp_file:
            stockval_df = parse_stock_validation_csv(tiktok_inv_file.read())
            sp_lookup = parse_tiktok_batch_edit(tiktok_mp_file.read())
            df = build_marketplace_df(stockval_df, sp_lookup, "TikTok")
            marketplace_data["TikTok"] = apply_product_names(df, name_lookup)

        if zalora_inv_file and zalora_mp_file:
            stockval_df = parse_stock_validation_csv(zalora_inv_file.read())
            sp_lookup = parse_zalora_stock_file(zalora_mp_file.read())
            status_lookup = parse_zalora_status_file(zalora_status_file.read()) if zalora_status_file else None
            df = build_marketplace_df(stockval_df, sp_lookup, "Zalora", status_lookup)
            marketplace_data["Zalora"] = apply_product_names(df, name_lookup)

        shopify_sp_lookup = None
        if shopify_mp_file:
            shopify_sp_lookup = parse_shopify_export(shopify_mp_file.read(), shopify_mp_file.name)
            if not shopify_sp_lookup:
                st.warning("Shopify MP file uploaded but no SKU/Available columns could be detected.")

        if shopify_inv_file and shopify_sp_lookup:
            stockval_df = parse_stock_validation_csv(shopify_inv_file.read())
            df = build_marketplace_df(stockval_df, shopify_sp_lookup, "Shopify")
            marketplace_data["Shopify"] = apply_product_names(df, name_lookup)

        soh_lookup = None
        if soh_file:
            soh_lookup = parse_soh(soh_file.read())

        if dtc_inv_file and soh_lookup:
            stockval_df = parse_stock_validation_csv(dtc_inv_file.read())
            df = build_marketplace_df(stockval_df, soh_lookup, "DTC")
            marketplace_data["DTC"] = apply_product_master_check(df, name_lookup)
        elif dtc_inv_file and not soh_file:
            st.warning("DTC inventory file uploaded, but the SOH warehouse file is needed to validate against — skipping DTC.")

        warehouse_df = None
        warehouse_vs_soh_df = None
        if warehouse_report_file and soh_lookup:
            wh_lookup = parse_warehouse_report(warehouse_report_file.read(), warehouse_report_file.name)
            if not wh_lookup:
                st.warning("Warehouse report uploaded but no SKU/quantity columns could be detected — skipping.")
            else:
                warehouse_vs_soh_df = apply_product_master_check(build_warehouse_vs_soh_df(wh_lookup, soh_lookup), name_lookup)
        elif warehouse_report_file and not soh_file:
            st.warning("Warehouse report uploaded, but SOH is needed to compare against — skipping.")
        elif soh_lookup and not warehouse_report_file:
            rows = [
                {"Seller SKU": sku, "Expected Stock": qty, "SP_Quantity": None,
                 "Status": "Mismatch", "Remark": "NOT FOUND"}
                for sku, qty in soh_lookup.items()
            ]
            if rows:
                warehouse_df = apply_product_master_check(pd.DataFrame(rows), name_lookup)
                st.info(
                    "SOH file loaded without a Warehouse report — showing warehouse stock as reference only "
                    "(no comparison available for this tab)."
                )

        mp_vs_pm_df = None
        if mp_report_file and name_lookup:
            mp_df = parse_mp_report(mp_report_file.read(), mp_report_file.name)
            if mp_df.empty:
                st.warning("MP Report file uploaded but no SKU column could be detected — skipping.")
            else:
                mp_vs_pm_df = build_mp_vs_product_master_df(mp_df, name_lookup)
        elif mp_report_file and not name_lookup:
            st.warning("MP Report file uploaded, but Product Master is needed to compare against — skipping.")

        pairs_to_check = [
            ("Lazada", lazada_mp_file, lazada_inv_file),
            ("Shopee", shopee_mp_file, tc_shopee_inv_file),
            ("Tiktok", tiktok_mp_file, tiktok_inv_file),
            ("Zalora", zalora_mp_file, zalora_inv_file),
            ("Shopify", shopify_mp_file, shopify_inv_file),
        ]
        missing_pairs = []
        for name, mp_f, inv_f in pairs_to_check:
            if mp_f and not inv_f:
                missing_pairs.append(f"{name} MP file uploaded without its inventory file")
            if inv_f and not mp_f:
                missing_pairs.append(f"{name} inventory file uploaded without its MP file")
        if missing_pairs:
            st.warning("Skipped incomplete pairs: " + "; ".join(missing_pairs))

        if (not marketplace_data and warehouse_df is None and mp_vs_pm_df is None
                and warehouse_vs_soh_df is None):
            st.error(
                "No complete pair was found. Each marketplace needs BOTH its "
                "MP file AND its inventory file uploaded (DTC needs the SOH warehouse file; "
                "MP Report needs Product Master; Warehouse report needs SOH)."
            )
        else:
            wb = build_workbook(marketplace_data, warehouse_df, brand_name, mp_vs_pm_df, warehouse_vs_soh_df)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            suffix = "Multi" if len(marketplace_data) > 1 else (list(marketplace_data)[0] if marketplace_data else "Report")
            filename = f"Stock_Validation_{suffix}_{datetime.now().strftime('%Y%m%d')}.xlsx"

            st.success("Validation complete!")
            metric_items = list(marketplace_data.items())
            if mp_vs_pm_df is not None:
                metric_items.append(("MP vs PM", mp_vs_pm_df))
            if warehouse_vs_soh_df is not None:
                metric_items.append(("Warehouse vs SOH", warehouse_vs_soh_df))
            if metric_items:
                cols = st.columns(len(metric_items))
                for i, (mp, df) in enumerate(metric_items):
                    with cols[i]:
                        mismatches = (df["Status"] == "Mismatch").sum()
                        st.metric(f"{mp} mismatches", int(mismatches), delta=None)

            st.download_button(
                "⬇️ Download Excel Report",
                data=buf,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("Upload each marketplace's MP file + inventory file into the matching card above to begin.")
